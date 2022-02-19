import openpyxl
import datetime
from openpyxl import Workbook, load_workbook

#wb = load_workbook("AHU Supply Return Daily Report.xlsx")

#"path of the workbook with file extension .xlsx"
ws = wb.active
ws['I7'].value = "Grevy Marcelin"
dateinput = datetime.date(July, 21, Wednesday)
print(dateinput)
ws['D8'].value = dateinput
wb.save("AHU Supply Return Daily Report.xlsx")

